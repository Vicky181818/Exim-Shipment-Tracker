"""
sync_excel.py — Read BL numbers from shipments.xlsx, scrape each carrier,
and write POD / ATA / Status back into the same file.
"""

import copy as _copy
import logging
import os
import re
import threading
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

EXCEL_FILE = "shipments.xlsx"

COLS   = ["BL No", "Carrier", "POD", "ATA / ETA", "Status"]
IDX    = {c: i + 1 for i, c in enumerate(COLS)}
WIDTHS = [22, 16, 28, 18, 36]

_BL_PREFIX = {
    # Hapag-Lloyd
    "HLCU": "HAPAG-LLOYD", "HLBU": "HAPAG-LLOYD", "HAMU": "HAPAG-LLOYD",
    "HLXU": "HAPAG-LLOYD", "BMOU": "HAPAG-LLOYD", "SEGU": "HAPAG-LLOYD",
    # Maersk
    "MAEU": "MAERSK",      "MRKU": "MAERSK",       "MSKU": "MAERSK",
    # MSC
    "MSCU": "MSC",         "MEDU": "MSC",           "MSDU": "MSC",         "TCKU": "MSC",
    # CMA CGM — container codes + regional BL formats (agency/port offices)
    "CMAU": "CMA CGM",     "CGMU": "CMA CGM",       "APHU": "CMA CGM",
    "LPL1": "CMA CGM",     "IBC1": "CMA CGM",       "SNG0": "CMA CGM",
    "DBA0": "CMA CGM",     "SYD6": "CMA CGM",       "RJK0": "CMA CGM",
    # COSCO
    "COSU": "COSCO",       "CBHU": "COSCO",         "CSNU": "COSCO",
    # Interasia — container codes and A02G-style booking BL numbers
    "IALU": "INTERASIA",   "IASU": "INTERASIA",
    "A02G": "INTERASIA",   "A02H": "INTERASIA",     "A02F": "INTERASIA",
    "A03G": "INTERASIA",   "A04G": "INTERASIA",
    # ONE LINE
    "ONEY": "ONE LINE",    "ONEU": "ONE LINE",       "CPTG": "ONE LINE",
    # HMM
    "HDMU": "HMM",         "HMMU": "HMM",            "BCNA": "HMM",         "HEMU": "HMM",
    # OOCL
    "OOLU": "OOCL",        "OOCU": "OOCL",           "ORCU": "OOCL",
    # KMTC — container code AND BL-number prefix (KMTCSIN..., KMTCPUS..., etc.)
    "KMTU": "KMTC",        "KMTC": "KMTC",
    # PIL
    "PCIU": "PIL",         "PILW": "PIL",
    # Trans Lines Shipping (TRLSINNSA... format)
    "TRLS": "TRANS LINE",
    # Blue Water Lines (JKT2... format)
    "JKT2": "BLUE WATER",
}

_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_FILL_RED    = PatternFill("solid", fgColor="FFC7CE")

_thin   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Scraping is sequential: one BL at a time, in sheet order. HTTP carriers
# take ~1-2s each; Selenium carriers reuse the API's pre-warmed Chrome pool.
# Parallel scraping was removed after COSCO 403-banned the IP for burst
# traffic — sequential is also what carrier sites expect from a real user.

_lock      = threading.Lock()   # guards _last_sync (read by /api/excel/status)
_last_sync: dict = {"running": False, "time": None, "pending": 0, "done": 0,
                    "total": 0, "ok": 0, "errors": 0, "message": "Never run"}

_MONTHS = ["January","February","March","April","May","June",
           "July","August","September","October","November","December"]


def _fmt_date(s: str) -> str:
    """'2026-07-06 02:00' → '6 July 2026 02:00', preserving 'ETA: ' prefix."""
    if not s:
        return ""
    prefix, rest = ("ETA: ", s[5:]) if s.startswith("ETA: ") else ("", s)
    try:
        dt = datetime.strptime(rest.strip()[:16], "%Y-%m-%d %H:%M")
        return f"{prefix}{dt.day} {_MONTHS[dt.month-1]} {dt.year} {dt.strftime('%H:%M')}"
    except Exception:
        return s

# Transshipment regex — catches "T/S", "T.S.", "TRANSSHIP" in event lines
_TS_RE = re.compile(r'\bT/S\b|T\.S\.|TRANSSHIP', re.I)


def _detect_carrier(bl: str) -> str:
    if bl.isdigit():
        return "MAERSK"
    return _BL_PREFIX.get(bl[:4].upper(), "")


def _clean_date(s: str) -> str:
    """Strip timezone suffix (IST, UTC…) and seconds, return YYYY-MM-DD HH:MM."""
    if not s:
        return ""
    s = re.sub(r'\s+[A-Z]{2,5}$', '', s.strip())  # remove trailing tz like "IST"
    return s[:16].strip()


def _determine_status(raw_st: str, ata: str, atd: str, pod: str) -> str:
    try:
        lines = [l for l in (raw_st or "").splitlines()
                 if l.strip() and not l.startswith("Status")]

        if lines:
            last = lines[-1]
            lu   = last.upper()

            # Unambiguous delivery signals
            if any(k in lu for k in ("TRANSPORT COMPLETED", "FULL TO CONSIGNEE",
                                     "DELIVERED TO CONSIGNEE", "DEHIRE", "DE-HIRE")):
                return "Delivered"

            # Container returned empty (empty return / received / in depot /
            # gate-in / de-hired) = the shipment is complete → Delivered.
            # Exclude origin empty-out events (pickup / release to shipper).
            if "EMPTY" in lu and not any(k in lu for k in
                    ("PICK", "RELEASE", "TO SHIPPER", "DISPATCH", "EMPTY OUT")):
                if any(k in lu for k in ("RETURN", "RECEIV", "DEPOT",
                                         "GATE IN", "DEHIRE", "DE-HIRE", "IN YARD")):
                    return "Delivered"

            # "DELIVERED" (past tense) = confirmed delivery
            # "DELIVER" alone is too broad — catches "Delivery Order Issued",
            # "Released for Delivery", "Delivery Instruction", etc.
            if "DELIVERED" in lu and not "UNDELIVERED" in lu:
                return "Delivered"

            # Gate Out: only counts as Delivered when it's at the POD.
            # Gate Out at origin (container leaving to vessel) = still In Transit.
            if any(k in lu for k in ("GATE OUT", "GATED OUT", "FULL OUT")):
                if pod:
                    pod_words = [w for w in re.sub(r'[,./()-]', ' ', pod.upper()).split()
                                 if len(w) >= 5]
                    if pod_words and not any(w in lu for w in pod_words):
                        return "In Transit"  # Gate out at non-POD = still moving
                return "Delivered"

            if _TS_RE.search(lu):
                return "In Transit"

            if any(k in lu for k in ("DISCHARG", "ARRIV", "GATED IN", "GATE IN",
                                     "RELEASED FOR DELIVERY", "AVAILABLE FOR DELIVERY",
                                     "AVAILABLE AT TERMINAL", "IMPORT RECEIVED")):
                if pod:
                    pod_words = [w for w in re.sub(r'[,./()-]', ' ', pod.upper()).split()
                                 if len(w) >= 5]
                    if pod_words and not any(w in lu for w in pod_words):
                        return "In Transit"
                return "Arrived / Discharged"

            if any(k in lu for k in ("LOAD", "DEPART", "SAIL", "TRANSIT", "VESSEL")):
                return "In Transit"

            # Unclassified event, but the shipment is clearly moving. Keep the
            # status clean — never leak raw carrier wording into the sheet.
            return "In Transit"

        if ata:
            try:
                ata_dt = datetime.strptime(_clean_date(ata)[:10], "%Y-%m-%d")
                if ata_dt.date() < datetime.now().date():
                    return "Arrived / Discharged"
                return "In Transit"
            except Exception:
                pass
        if atd:
            return "In Transit"
        return "Booking / Pre-Departure"
    except Exception:
        return "Unknown"


def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    ws.row_dimensions[1].height = 30

    for i, (col_name, width) in enumerate(zip(COLS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _ALIGN_C
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)
    log.info("[EXCEL] Created %s", EXCEL_FILE)


def ensure_excel():
    if not os.path.exists(EXCEL_FILE):
        create_excel()


def _status_fill(status: str) -> PatternFill:
    s = (status or "").upper()
    if any(k in s for k in ("DELIVERED", "ARRIVED", "DISCHARGED")):
        return _FILL_GREEN
    if "ERROR" in s or "NOT FOUND" in s:
        return _FILL_RED
    return _FILL_YELLOW


def _write_row(ws, row: int, pod: str, ata: str, status: str):
    try:
        fill = _status_fill(status)
        for col_name, value in [("POD", pod), ("ATA / ETA", ata), ("Status", status)]:
            c = ws.cell(row=row, column=IDX[col_name], value=str(value) if value else "")
            c.font      = Font(name="Calibri", size=10)
            c.fill      = fill
            c.alignment = _ALIGN_L
            c.border    = _BORDER
    except Exception as e:
        log.error("[SYNC] _write_row(%d) failed: %s", row, e)
        raise  # let caller handle


def _write_status_only(ws, row: int, status: str):
    """Update only the Status cell — used when a scrape fails so a transient
    error (rate limit, timeout) never wipes the existing POD / ATA data."""
    c = ws.cell(row=row, column=IDX["Status"], value=status)
    c.font      = Font(name="Calibri", size=10)
    c.fill      = _status_fill(status)
    c.alignment = _ALIGN_L
    c.border    = _BORDER


def _save(wb) -> bool:
    """Save workbook; return False if file is locked (e.g. open in Excel)."""
    try:
        wb.save(EXCEL_FILE)
        return True
    except PermissionError:
        log.warning("[SYNC] Cannot save — file is open in another program (Excel?)")
        return False
    except Exception as e:
        log.warning("[SYNC] Save failed: %s", e)
        return False


def run_sync(scrape_fn) -> dict:
    with _lock:
        _last_sync.update({"running": True, "message": "Starting…",
                           "total": 0, "ok": 0, "errors": 0, "done": 0, "pending": 0})

    cnt = {"done": 0, "ok": 0, "errors": 0}

    try:
        ensure_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        first_cell = str(ws.cell(1, 1).value or "").strip()
        has_header = first_cell.upper() in ("BL NO", "BL", "B/L NO", "B/L")
        data_start = 2 if has_header else 1

        header_row  = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        bl_col      = header_row.get("BL No",  IDX["BL No"])
        carrier_col = header_row.get("Carrier", IDX["Carrier"])

        # ── Phase 1: Read all rows sequentially, detect carriers ─────────────
        seen_bls     = {}   # bl -> first_row_number
        unique_tasks = []   # (row, bl, carrier) — one scrape each
        dup_rows     = []   # (row, bl)          — copy from first occurrence

        for row in range(data_start, ws.max_row + 1):
            try:
                bl = str(ws.cell(row, bl_col).value or "").strip()
                if not bl:
                    continue

                if bl in seen_bls:
                    dup_rows.append((row, bl))
                    continue

                carrier = str(ws.cell(row, carrier_col).value or "").strip()
                if not carrier:
                    carrier = _detect_carrier(bl)
                    if carrier:
                        try:
                            cell = ws.cell(row, carrier_col)
                            cell.value     = carrier
                            cell.font      = Font(name="Calibri", size=10)
                            cell.alignment = _ALIGN_L
                            cell.border    = _BORDER
                        except Exception:
                            pass

                seen_bls[bl] = row
                unique_tasks.append((row, bl, carrier))
            except Exception as e:
                log.error("[SYNC] Setup row %d error: %s", row, e)

        pending = len(unique_tasks)
        with _lock:
            _last_sync.update({"pending": pending, "done": 0})

        _save(wb)  # persist auto-detected carriers before parallel scraping

        # ── Phase 2: Scrape unique BLs one at a time ─────────────────────────
        def _scrape_row(row: int, bl: str, carrier: str):
            if not carrier:
                return row, bl, "", "", "Error: carrier not detected", False
            log.info("[SYNC] %s  carrier=%s", bl, carrier)
            try:
                data   = scrape_fn(bl, carrier)
                pod    = data.get("POD") or ""
                ata    = _clean_date(data.get("ATA") or "")
                atd    = _clean_date(data.get("ATD") or "")
                raw_st = data.get("Latest Status") or ""
                error  = data.get("error") or ""

                if error:
                    return row, bl, pod, "", f"Error: {error}", False

                status = _determine_status(raw_st, ata, atd, pod)
                if status == "In Transit" and ata:
                    try:
                        if datetime.strptime(ata[:10], "%Y-%m-%d").date() >= datetime.now().date():
                            ata = f"ETA: {ata}"
                    except Exception:
                        pass
                return row, bl, pod, _fmt_date(ata), status, True

            except Exception as e:
                log.error("[SYNC] %s scrape error: %s", bl, e)
                return row, bl, "", "", f"Error: {str(e)[:70]}", False

        for r, b, c in unique_tasks:
            row, bl, pod, ata, status, ok = _scrape_row(r, b, c)

            cnt["done"] += 1
            if ok:
                cnt["ok"] += 1
            else:
                cnt["errors"] += 1
            try:
                if ok:
                    _write_row(ws, row, pod, ata, status)
                else:
                    _write_status_only(ws, row, status)
                _save(wb)
            except Exception as we:
                log.error("[SYNC] Write row %d failed: %s", row, we)
            with _lock:
                _last_sync.update({
                    "running": True, "pending": pending,
                    "done": cnt["done"], "total": cnt["done"],
                    "ok": cnt["ok"], "errors": cnt["errors"],
                    "message": f"Syncing {bl}…",
                })

        # ── Phase 3: Copy results to duplicate rows ───────────────────────────
        for row, bl in dup_rows:
            try:
                first = seen_bls[bl]
                for col in (IDX["POD"], IDX["ATA / ETA"], IDX["Status"], IDX["Carrier"]):
                    src = ws.cell(first, col)
                    dst = ws.cell(row, col)
                    dst.value     = src.value
                    dst.font      = _copy.copy(src.font)
                    dst.fill      = _copy.copy(src.fill)
                    dst.alignment = _copy.copy(src.alignment)
                    dst.border    = _copy.copy(src.border)
            except Exception as de:
                log.error("[SYNC] Dup row %d failed: %s", row, de)
                try:
                    _write_row(ws, row, "", "", "Error: duplicate copy failed")
                except Exception:
                    pass
                cnt["errors"] += 1

        _save(wb)

    except Exception as fatal_e:
        log.error("[SYNC] Fatal: %s", fatal_e, exc_info=True)
        cnt["errors"] += 1

    total = cnt["done"]
    result = {
        "running": False,
        "time":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        "pending": total, "done": total, "total": total,
        "ok":      cnt["ok"],
        "errors":  cnt["errors"],
        "message": f"Done — {cnt['ok']}/{total} synced, {cnt['errors']} errors",
    }
    with _lock:
        _last_sync.update(result)
    log.info("[SYNC] %s", result["message"])
    return result


def get_status() -> dict:
    return dict(_last_sync)
