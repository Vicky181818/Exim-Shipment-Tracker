import time
import logging
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

def make_driver(headless: bool = False, use_profile: bool = False, profile_dir: str = None):
    """
    Create a Chrome WebDriver.

    headless    — run without a visible window (some carriers break in headless mode)
    use_profile — load the user's real Chrome profile so existing login sessions
                  (e.g. ONE LINE ecomm portal) are inherited.
                  ⚠ Chrome must be fully closed before calling this, otherwise
                  Chrome will refuse to open a second instance on the same profile.
    profile_dir — path to a dedicated Chrome profile directory (auto-created on first run).
                  Use this instead of use_profile to avoid conflicts with the user's main Chrome.
                  Cloudflare clearance cookies stored here survive between runs.
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    opts = webdriver.ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    if profile_dir:
        # Dedicated profile directory — Chrome creates it on first launch.
        # Cookies (including CF clearance) persist across runs.
        os.makedirs(profile_dir, exist_ok=True)
        opts.add_argument(f"--user-data-dir={profile_dir}")
    elif use_profile:
        # Windows default Chrome profile path
        profile_path = os.path.join(
            os.environ.get("LOCALAPPDATA", ""),
            "Google", "Chrome", "User Data"
        )
        if os.path.isdir(profile_path):
            opts.add_argument(f"--user-data-dir={profile_path}")
            opts.add_argument("--profile-directory=Default")
        else:
            log.warning("Chrome profile not found at %s — launching without profile", profile_path)

    cd_path = ChromeDriverManager().install()
    os.system(f"codesign -f -s - '{cd_path}' 2>/dev/null")

    service = Service(cd_path)
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    driver.set_page_load_timeout(30)
    return driver

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

from scrapers.maersk    import scrape as maersk_scrape
from scrapers.msc       import scrape as msc_scrape
from scrapers.cma       import scrape as cma_scrape
from scrapers.cosco     import scrape as cosco_scrape
from scrapers.hapag     import scrape as hapag_scrape
from scrapers.interasia import scrape as interasia_scrape
from scrapers.hmm       import scrape as hmm_scrape
from scrapers.one_line  import scrape as one_line_scrape
from scrapers.pil       import scrape as pil_scrape
from scrapers.kmtc      import scrape as kmtc_scrape
from scrapers.oocl      import scrape as oocl_scrape
from scrapers.transliner import scrape as transliner_scrape


EXCEL_FILE   = "tracking_2.xlsx"
SHEET_INPUT  = "BL Input"
SHEET_OUTPUT = "Tracking Results"
HEADLESS     = False
DELAY        = 2
LOG_DIR      = "logs"

OUTPUT_COLS = [
    "BL No", "Shipping Line", "POL", "POD",
    "Container No", "Vessel", "ATA", "FND",
    "Latest Status", "Last Updated",
]

DARK_BLUE  = "1B3A6B"
GREEN_BG   = "D5F5E3"
YELLOW_BG  = "FEF9E7"
RED_BG     = "FADBD8"
ALT_ROW    = "F0F4FA"

os.makedirs(LOG_DIR, exist_ok=True)
log_file = os.path.join(LOG_DIR, datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(),
    ]
)
log = logging.getLogger(__name__)

SCRAPER_MAP = {
    # ── Working carriers ────────────────────────────────────────────────
    "MAERSK":      maersk_scrape,
    "MSC":         msc_scrape,
    "CMA CGM":     cma_scrape,
    "CMA":         cma_scrape,
    "COSCO":       cosco_scrape,
    "HAPAG-LLOYD": hapag_scrape,
    "HAPAG LLOYD": hapag_scrape,
    "HAPAG":       hapag_scrape,
    "INTERASIA":   interasia_scrape,
    "INTER ASIA":  interasia_scrape,

    # ── New carriers ────────────────────────────────────────────────────
    # HMM (Hyundai Merchant Marine) — Selenium, hmm21.com
    "HMM LINE":    hmm_scrape,
    "HMM":         hmm_scrape,

    # ONE LINE — Selenium, ecomm.one-line.com
    "ONE LINE":    one_line_scrape,
    "ONE":         one_line_scrape,

    # PIL (Pacific International Lines) — HTTP API, no browser needed
    "PIL":         pil_scrape,

    # KMTC — Selenium, koreascl.com
    "KMTC":        kmtc_scrape,

    # OOCL — Selenium, oocl.com
    "OOCL":        oocl_scrape,

    # Transliner (Transliner Pte Ltd) — HTTP API (Tigris), no browser needed
    "TRANSLINER":  transliner_scrape,
    "TRANS LINE":  transliner_scrape,
    "TRANS LINES": transliner_scrape,
    "TRANSLINE":   transliner_scrape,
}

def get_scraper(line_raw: str):
    key = str(line_raw).strip().upper()
    return SCRAPER_MAP.get(key)



def ensure_sheets(wb: openpyxl.Workbook):
    if SHEET_INPUT not in wb.sheetnames:
        ws1 = wb.create_sheet(SHEET_INPUT, 0)
        _write_header(ws1, ["BL No", "Shipping Line"], DARK_BLUE)
        log.info("Created sheet: %s", SHEET_INPUT)

    if SHEET_OUTPUT not in wb.sheetnames:
        ws2 = wb.create_sheet(SHEET_OUTPUT)
        _write_header(ws2, OUTPUT_COLS, DARK_BLUE)
        ws2.freeze_panes = "A2"
        log.info("Created sheet: %s", SHEET_OUTPUT)

def _write_header(ws, cols, color):
    hf = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", start_color=color)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = fill
        c.alignment = center
        c.border = border

    ws.row_dimensions[1].height = 22

def read_sheet1(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb[SHEET_INPUT]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bl   = str(row[0]).strip() if row[0] else ""
        line = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if bl and bl != "nan" and line and line != "nan":
            rows.append({"bl": bl, "line": line})
    return rows

def read_sheet2_index(wb: openpyxl.Workbook) -> dict:
    ws = wb[SHEET_OUTPUT]
    index = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        bl_cell = row[0]
        if bl_cell.value:
            index[str(bl_cell.value).strip()] = bl_cell.row
    return index

def write_result(ws, row_num: int, data: dict, is_new: bool):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    df = Font(name="Arial", size=9)
    left  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")

    status = str(data.get("Latest Status", "")).lower()
    if any(w in status for w in ("empty return", "delivered", "gate out", "returned to carrier", "empty available", "return")):
        status_fill = PatternFill("solid", start_color=GREEN_BG)
        status_font = Font(name="Arial", size=9, bold=True, color="1E8449")
    elif any(w in status for w in ("error", "manual", "not found", "timeout")):
        status_fill = PatternFill("solid", start_color=RED_BG)
        status_font = Font(name="Arial", size=9, bold=True, color="C0392B")
    else:
        status_fill = PatternFill("solid", start_color=YELLOW_BG)
        status_font = Font(name="Arial", size=9, bold=True, color="7D6608")

    alt_fill = PatternFill("solid", start_color=ALT_ROW) if row_num % 2 == 0 else None

    for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
        val  = data.get(col_name, "")
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = border

        if col_name == "Latest Status":
            cell.fill = status_fill
            cell.font = status_font
            cell.alignment = left
        else:
            cell.font = df
            cell.alignment = center if col_idx <= 2 else left
            if alt_fill:
                cell.fill = alt_fill

    ws.row_dimensions[row_num].height = 18

def update_col_widths(ws):
    widths = {
        "BL No": 18, "Shipping Line": 15, "POL": 20, "POD": 20,
        "Container No": 18, "Vessel": 22, "ATA": 20,
        "Latest Status": 28, "Last Updated": 18,"FND": 20,
    }
    for i, col in enumerate(OUTPUT_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)

def main():
    log.info("=" * 55)
    log.info("  SHIPPING TRACKER STARTED")
    log.info("=" * 55)

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        log.info("Loaded: %s", EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        log.info("Created new workbook: %s", EXCEL_FILE)

    ensure_sheets(wb)
    wb.save(EXCEL_FILE)

    inputs = read_sheet1(wb)
    if not inputs:
        log.warning("Sheet 1 is empty — add BL No + Shipping Line rows and re-run.")
        return

    log.info("Found %d BL rows in Sheet 1", len(inputs))

    log.info("Starting Chrome...")
    driver = make_driver()

    results = []
    try:
        for i, item in enumerate(inputs, 1):
            bl   = item["bl"]
            line = item["line"]
            scraper = get_scraper(line)

            log.info("[%d/%d] %s | %s", i, len(inputs), line, bl)

            if not scraper:
                log.warning("  No tracker for line: %s — skipping", line)
                data = {col: "" for col in OUTPUT_COLS}
                data["BL No"]        = bl
                data["Shipping Line"] = line
                data["Latest Status"] = f"Carrier not supported: {line}"
                data["Last Updated"]  = datetime.now().strftime("%Y-%m-%d %H:%M")
            else:
                try:
                    data = scraper(driver, bl)

                    print("\n========================")
                    print("BL:", bl)
                    print("SCRAPER RESULT:")
                    print(data)
                    print("========================\n")

                    useful_fields = ["POL", "POD", "Container No", "Vessel", "ATA", "Latest Status"]
                    if all(not str(data.get(f, "")).strip() for f in useful_fields):
                        raise ValueError("Scraper returned empty data")

                    data["BL No"] = bl
                    data["Shipping Line"] = line
                    data["Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")

                    log.info("  ✅ %s", data.get("Latest Status", "")[:60])

                except Exception as e:
                    log.error("  ❌ Error scraping %s / %s: %s", line, bl, e)
                    data = {col: "" for col in OUTPUT_COLS}
                    data["BL No"] = bl
                    data["Shipping Line"] = line
                    data["Latest Status"] = f"Error: {str(e)[:80]}"
                    data["Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")

            results.append(data)
            time.sleep(DELAY)

    finally:
        driver.quit()
        log.info("Browser closed.")

    wb2 = openpyxl.load_workbook(EXCEL_FILE)
    ensure_sheets(wb2)
    ws2     = wb2[SHEET_OUTPUT]
    bl_index = read_sheet2_index(wb2)

    for data in results:
        bl = data["BL No"]
        if bl in bl_index:
            row_num  = bl_index[bl]
            is_new   = False
            log.info("Updated row %d → %s", row_num, bl)
        else:
            row_num  = ws2.max_row + 1
            is_new   = True
            log.info("Appended row %d → %s", row_num, bl)

        write_result(ws2, row_num, data, is_new)

    update_col_widths(ws2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLS))}1"

    wb2.save(EXCEL_FILE)
    log.info("Saved: %s", EXCEL_FILE)

    ok = sum(
        1 for d in results
        if any(
            str(d.get(k, "")).strip()
            for k in [
                "POL",
                "POD",
                "Container No",
                "Vessel",
                "ATA",
                "Latest Status",
            ]
        )
        and "error" not in str(d.get("Latest Status", "")).lower()
    )

    errors = len(results) - ok
    log.info("=" * 55)
    log.info("DONE — %d scraped OK, %d errors", ok, errors)
    log.info("Log saved: %s", log_file)
    log.info("=" * 55)

if __name__ == "__main__":
    main()
